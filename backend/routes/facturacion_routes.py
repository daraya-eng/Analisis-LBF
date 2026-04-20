"""
Adjudicado vs Facturado — Licitaciones vigentes con cruce BI_TOTAL_FACTURA.
Solo facturación con TIPO_OC='LICITACION'.
"""
import datetime
import io
import json
import os
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from auth import get_current_user
from db import get_conn, hoy, filtro_guias
from cache import mem_get, mem_set, _mem_cache

_NOTAS_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "notas_licitaciones.json")


def _load_notas() -> dict:
    if os.path.exists(_NOTAS_PATH):
        with open(_NOTAS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_notas(notas: dict):
    os.makedirs(os.path.dirname(_NOTAS_PATH), exist_ok=True)
    with open(_NOTAS_PATH, "w", encoding="utf-8") as f:
        json.dump(notas, f, ensure_ascii=False, indent=2)

router = APIRouter()


def _fmt_fecha(val) -> str:
    """Convert date to dd-mm-yyyy string."""
    if not val:
        return ""
    s = str(val).strip()
    if len(s) >= 10 and "-" in s:
        parts = s[:10].split("-")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return s


def _load_facturacion() -> dict:
    h = hoy()
    _ANO, _MES, _HOY = h["ano"], h["mes"], h["hoy"]
    _MES_NOMBRE, _MES_PREFIX = h["mes_nombre"], h["mes_prefix"]
    _FG = filtro_guias()

    conn = get_conn()
    cur = conn.cursor()

    # ── KAM por RUT (último KAM registrado en BI_TOTAL_FACTURA) ──
    cur.execute(f"""
        SELECT RUT, MAX(KAM) AS KAM
        FROM BI_TOTAL_FACTURA
        WHERE KAM IS NOT NULL AND KAM != ''
          AND {_FG}
        GROUP BY RUT
    """)
    kam_map = {}
    for r in cur.fetchall():
        rut = str(r[0] or "").strip()
        if rut:
            kam_map[rut] = str(r[1] or "").strip()

    # Fallback: FFVV_ZONA de licitaciones para RUTs sin KAM en BI
    cur.execute("""
        SELECT rut_cliente, MAX(FFVV_ZONA) AS zona
        FROM vw_LICITACIONES_CATEGORIZADAS
        WHERE EsLBF = 1 AND FFVV_ZONA IS NOT NULL AND FFVV_ZONA != ''
        GROUP BY rut_cliente
    """)
    zona_map = {}
    for r in cur.fetchall():
        rut = str(r[0] or "").strip()
        if rut and rut not in kam_map:
            zona_map[rut] = str(r[1] or "").strip()

    # ── Licitaciones vigentes LBF, ordenadas por fecha_termino ASC ──
    cur.execute(f"""
        WITH adj AS (
            SELECT licitacion, MAX(rut_cliente) AS rut, MAX(nombre_cliente) AS nombre,
                SUM(TRY_CAST(monto_licitacion AS bigint)) AS monto_adjudicado,
                MAX(fecha_termino) AS fecha_termino,
                MAX(fecha_inicio) AS fecha_inicio
            FROM vw_LICITACIONES_CATEGORIZADAS
            WHERE EsLBF = 1 AND estado = 'Adjudicado'
              AND TRY_CAST(fecha_termino AS date) >= '{_HOY}'
            GROUP BY licitacion
        ),
        fac AS (
            SELECT LICITACION, SUM(CAST(VENTA AS float)) AS facturado
            FROM BI_TOTAL_FACTURA
            WHERE TIPO_OC = 'LICITACION'
              AND {_FG}
            GROUP BY LICITACION
        )
        SELECT adj.licitacion, adj.rut, adj.nombre, adj.monto_adjudicado,
               adj.fecha_inicio, adj.fecha_termino,
               COALESCE(fac.facturado, 0) AS facturado,
               DATEDIFF(day, '{_HOY}', adj.fecha_termino) AS dias_restantes
        FROM adj
        LEFT JOIN fac ON adj.licitacion = fac.LICITACION
        ORDER BY adj.fecha_termino ASC
    """)
    licitaciones = []
    total_adj = 0
    total_fac = 0
    urgentes = 0
    # Urgentes reales: vencen en el mes en curso y cumplimiento < 100%
    urgentes_reales = []
    for r in cur.fetchall():
        adj_val = int(r[3] or 0)
        fac_val = round(float(r[6] or 0))
        dias = int(r[7] or 0)
        pct = round(fac_val / adj_val * 100, 1) if adj_val > 0 else 0
        total_adj += adj_val
        total_fac += fac_val
        if dias <= 30:
            urgentes += 1
        sem = "red" if dias <= 30 else ("yellow" if dias <= 90 else "green")
        rut_str = str(r[1] or "").strip()
        kam = kam_map.get(rut_str) or zona_map.get(rut_str, "Sin KAM")
        row = {
            "licitacion": str(r[0] or "").strip(),
            "rut": rut_str,
            "nombre": str(r[2] or "").strip(),
            "adjudicado": adj_val,
            "fecha_inicio": _fmt_fecha(r[4]),
            "fecha_termino": _fmt_fecha(r[5]),
            "facturado": fac_val,
            "cumplimiento": pct,
            "dias_restantes": dias,
            "semaforo": sem,
            "kam": kam,
        }
        licitaciones.append(row)
        # Urgentes reales: vencen en el mes en curso y NO están al 100%
        ft_raw = str(r[5] or "")
        if ft_raw[:7] == _MES_PREFIX and pct < 100:
            urgentes_reales.append(row)

    # KPIs urgentes reales
    ur_adj = sum(u["adjudicado"] for u in urgentes_reales)
    ur_fac = sum(u["facturado"] for u in urgentes_reales)
    ur_gap = ur_adj - ur_fac

    kpis = {
        "total_vigentes": len(licitaciones),
        "total_adjudicado": total_adj,
        "total_facturado": round(total_fac),
        "cumplimiento": round(total_fac / total_adj * 100, 1) if total_adj > 0 else 0,
        "gap": total_adj - round(total_fac),
        "urgentes_30d": urgentes,
        "urgentes_reales": len(urgentes_reales),
        "urgentes_reales_gap": ur_gap,
        "urgentes_reales_adj": ur_adj,
        "urgentes_reales_fac": ur_fac,
    }

    # ── Por cliente (agrupado) ──
    cur.execute(f"""
        WITH adj AS (
            SELECT rut_cliente, MAX(nombre_cliente) AS nombre,
                SUM(TRY_CAST(monto_licitacion AS bigint)) AS monto_adjudicado,
                COUNT(DISTINCT licitacion) AS n_lic,
                MIN(DATEDIFF(day, '{_HOY}', fecha_termino)) AS min_dias
            FROM vw_LICITACIONES_CATEGORIZADAS
            WHERE EsLBF = 1 AND estado = 'Adjudicado'
              AND TRY_CAST(fecha_termino AS date) >= '{_HOY}'
            GROUP BY rut_cliente
        ),
        fac AS (
            SELECT RUT, SUM(CAST(VENTA AS float)) AS facturado
            FROM BI_TOTAL_FACTURA
            WHERE TIPO_OC = 'LICITACION'
              AND {_FG}
            GROUP BY RUT
        )
        SELECT adj.rut_cliente, adj.nombre, adj.monto_adjudicado, adj.n_lic,
               adj.min_dias, COALESCE(fac.facturado, 0) AS facturado
        FROM adj
        LEFT JOIN fac ON adj.rut_cliente = fac.RUT
        ORDER BY adj.monto_adjudicado DESC
    """)
    clientes = []
    for r in cur.fetchall():
        adj_val = int(r[2] or 0)
        fac_val = round(float(r[5] or 0))
        pct = round(fac_val / adj_val * 100, 1) if adj_val > 0 else 0
        dias = int(r[4] or 0)
        rut_str = str(r[0] or "").strip()
        clientes.append({
            "rut": rut_str,
            "nombre": str(r[1] or "").strip(),
            "adjudicado": adj_val,
            "n_licitaciones": r[3] or 0,
            "facturado": fac_val,
            "cumplimiento": pct,
            "dias_mas_pronto": dias,
            "semaforo": "red" if pct < 50 else ("yellow" if pct < 80 else "green"),
            "kam": kam_map.get(rut_str) or zona_map.get(rut_str, "Sin KAM"),
        })

    # ── Canales de venta (para contexto) ──
    cur.execute(f"""
        SELECT ISNULL(TIPO_OC, 'Otro') AS canal,
            SUM(CAST(VENTA AS float)) AS venta,
            COUNT(DISTINCT RUT) AS n_clientes
        FROM BI_TOTAL_FACTURA
        WHERE ANO = {_ANO}
          AND {_FG}
          AND TIPO_OC IS NOT NULL AND TIPO_OC != ''
        GROUP BY ISNULL(TIPO_OC, 'Otro')
        ORDER BY venta DESC
    """)
    canales = [{"canal": str(r[0] or "").strip(), "venta": round(float(r[1] or 0)),
                "n_clientes": r[2] or 0} for r in cur.fetchall()]

    # Ordenar licitaciones y clientes por cumplimiento ASC (peor primero)
    licitaciones.sort(key=lambda x: x["cumplimiento"])
    clientes.sort(key=lambda x: x["cumplimiento"])
    urgentes_reales.sort(key=lambda x: x["cumplimiento"])

    conn.close()

    # Adjuntar notas existentes a cada licitación
    notas = _load_notas()
    for l in licitaciones:
        n = notas.get(l["licitacion"])
        if n:
            l["nota"] = n
    for u in urgentes_reales:
        n = notas.get(u["licitacion"])
        if n:
            u["nota"] = n

    # Lista de KAMs únicos para filtro
    kams_set = sorted(set(l["kam"] for l in licitaciones))

    return {"kpis": kpis, "licitaciones": licitaciones, "clientes": clientes,
            "canales": canales, "urgentes_reales": urgentes_reales,
            "mes_nombre": _MES_NOMBRE, "kams": kams_set}


def _load_detalle_licitacion(licitacion: str) -> dict:
    """Detalle de una licitación: productos adjudicados + facturados."""
    _FG = filtro_guias()
    conn = get_conn()
    cur = conn.cursor()

    # Productos adjudicados en la licitación (desde vw_LICITACIONES)
    cur.execute("""
        SELECT descripcion_producto, DescripcionMaestro, Categoria,
               TRY_CAST(monto_licitacion AS bigint) AS monto,
               nombre_empresa, estado
        FROM vw_LICITACIONES_CATEGORIZADAS
        WHERE licitacion = ? AND EsLBF = 1
        ORDER BY TRY_CAST(monto_licitacion AS bigint) DESC
    """, (licitacion,))
    adjudicados = []
    for r in cur.fetchall():
        adjudicados.append({
            "producto_licitacion": str(r[0] or "").strip(),
            "producto_lbf": str(r[1] or "").strip(),
            "categoria": str(r[2] or "").strip(),
            "monto": int(r[3] or 0),
            "empresa": str(r[4] or "").strip(),
            "estado": str(r[5] or "").strip(),
        })

    # Facturación real por producto (desde BI_TOTAL_FACTURA), separado por DOC_CODE
    cur.execute(f"""
        SELECT CODIGO, DESCRIPCION, DOC_CODE,
               SUM(CAST(VENTA AS float)) AS venta,
               COUNT(*) AS n_docs,
               MAX(DIA) AS ultima_fecha
        FROM BI_TOTAL_FACTURA
        WHERE LICITACION = ? AND TIPO_OC = 'LICITACION'
          AND {_FG}
        GROUP BY CODIGO, DESCRIPCION, DOC_CODE
        ORDER BY CODIGO, DOC_CODE
    """, (licitacion,))
    facturados = []
    for r in cur.fetchall():
        doc = str(r[2] or "").strip()
        facturados.append({
            "codigo": str(r[0] or "").strip(),
            "descripcion": str(r[1] or "").strip(),
            "doc_code": doc,
            "tipo": "Nota Crédito" if doc in ("NT", "CM") else ("Guía" if doc == "GF" else "Factura"),
            "venta": round(float(r[3] or 0)),
            "n_docs": int(r[4] or 0),
            "ultima_fecha": str(r[5] or "")[:10],
        })

    conn.close()
    notas = _load_notas()
    nota = notas.get(licitacion)
    return {"licitacion": licitacion, "adjudicados": adjudicados, "facturados": facturados, "nota": nota}


@router.get("/")
async def get_facturacion(
    current_user: dict = Depends(get_current_user),
):
    try:
        ck = "facturacion_vigentes"
        cached = mem_get(ck)
        if cached:
            return cached
        data = _load_facturacion()
        mem_set(ck, data)
        return data
    except Exception as e:
        return {"error": str(e), "kpis": {}, "licitaciones": [], "clientes": [], "canales": []}


@router.get("/detalle")
async def get_detalle_licitacion(
    licitacion: str = Query(...),
    current_user: dict = Depends(get_current_user),
):
    try:
        ck = f"fac_det_{licitacion}"
        cached = mem_get(ck)
        if cached:
            return cached
        data = _load_detalle_licitacion(licitacion)
        mem_set(ck, data)
        return data
    except Exception as e:
        return {"error": str(e), "adjudicados": [], "facturados": []}


class DetalleBatchBody(BaseModel):
    licitaciones: List[str]


@router.post("/detalle-batch")
async def get_detalle_batch(
    body: DetalleBatchBody,
    current_user: dict = Depends(get_current_user),
):
    """Detalle de productos para múltiples licitaciones (para email)."""
    try:
        if not body.licitaciones:
            return {}
        _FG = filtro_guias()
        conn = get_conn()
        cur = conn.cursor()
        placeholders = ",".join(["?" for _ in body.licitaciones])

        # Adjudicados
        cur.execute(f"""
            SELECT licitacion, descripcion_producto, DescripcionMaestro,
                   TRY_CAST(monto_licitacion AS bigint) AS monto
            FROM vw_LICITACIONES_CATEGORIZADAS
            WHERE licitacion IN ({placeholders}) AND EsLBF = 1
            ORDER BY licitacion, TRY_CAST(monto_licitacion AS bigint) DESC
        """, body.licitaciones)
        adj_map: dict = {}
        for r in cur.fetchall():
            lic = str(r[0] or "").strip()
            if lic not in adj_map:
                adj_map[lic] = []
            adj_map[lic].append({
                "producto": str(r[1] or "").strip(),
                "producto_lbf": str(r[2] or "").strip(),
                "monto": int(r[3] or 0),
            })

        # Facturados
        cur.execute(f"""
            SELECT LICITACION, CODIGO, DESCRIPCION,
                   SUM(CAST(VENTA AS float)) AS venta
            FROM BI_TOTAL_FACTURA
            WHERE LICITACION IN ({placeholders}) AND TIPO_OC = 'LICITACION'
              AND {_FG}
            GROUP BY LICITACION, CODIGO, DESCRIPCION
            ORDER BY LICITACION, SUM(CAST(VENTA AS float)) DESC
        """, body.licitaciones)
        fac_map: dict = {}
        for r in cur.fetchall():
            lic = str(r[0] or "").strip()
            if lic not in fac_map:
                fac_map[lic] = []
            fac_map[lic].append({
                "codigo": str(r[1] or "").strip(),
                "descripcion": str(r[2] or "").strip(),
                "venta": round(float(r[3] or 0)),
            })

        conn.close()

        result = {}
        for lic in body.licitaciones:
            result[lic] = {
                "adjudicados": adj_map.get(lic, []),
                "facturados": fac_map.get(lic, []),
            }
        return result
    except Exception as e:
        return {"error": str(e)}


class ExcelEmailBody(BaseModel):
    licitaciones: List[str]
    kam: str


@router.post("/excel-detalle")
async def get_excel_detalle(
    body: ExcelEmailBody,
    current_user: dict = Depends(get_current_user),
):
    """Genera Excel con detalle de productos por licitación para un KAM."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _FG = filtro_guias()
    conn = get_conn()
    cur = conn.cursor()
    placeholders = ",".join(["?" for _ in body.licitaciones])

    # Datos de licitaciones (resumen)
    cur.execute(f"""
        SELECT l.licitacion, l.rut_cliente, l.nombre_cliente,
               SUM(TRY_CAST(l.monto_licitacion AS bigint)) AS adjudicado,
               MIN(l.fecha_inicio) AS fecha_inicio,
               MAX(l.fecha_termino) AS fecha_termino
        FROM vw_LICITACIONES_CATEGORIZADAS l
        WHERE l.licitacion IN ({placeholders}) AND l.EsLBF = 1
        GROUP BY l.licitacion, l.rut_cliente, l.nombre_cliente
        ORDER BY l.licitacion
    """, body.licitaciones)
    resumen = []
    for r in cur.fetchall():
        resumen.append({
            "licitacion": str(r[0] or "").strip(),
            "rut": str(r[1] or "").strip(),
            "cliente": str(r[2] or "").strip(),
            "adjudicado": int(r[3] or 0),
            "fecha_inicio": str(r[4] or "")[:10],
            "fecha_termino": str(r[5] or "")[:10],
        })

    # Adjudicados por licitación
    cur.execute(f"""
        SELECT licitacion, descripcion_producto, DescripcionMaestro,
               TRY_CAST(monto_licitacion AS bigint) AS monto, estado
        FROM vw_LICITACIONES_CATEGORIZADAS
        WHERE licitacion IN ({placeholders}) AND EsLBF = 1
              AND TRY_CAST(monto_licitacion AS bigint) > 0
        ORDER BY licitacion, TRY_CAST(monto_licitacion AS bigint) DESC
    """, body.licitaciones)
    adj_map: dict = {}
    for r in cur.fetchall():
        lic = str(r[0] or "").strip()
        if lic not in adj_map:
            adj_map[lic] = []
        adj_map[lic].append({
            "producto": str(r[1] or "").strip(),
            "producto_lbf": str(r[2] or "").strip(),
            "monto": int(r[3] or 0),
            "estado": str(r[4] or "").strip(),
        })

    # Facturados por licitación
    cur.execute(f"""
        SELECT LICITACION, CODIGO, DESCRIPCION, DOC_CODE,
               SUM(CAST(VENTA AS float)) AS venta,
               COUNT(*) AS n_docs,
               MAX(DIA) AS ultima_fecha
        FROM BI_TOTAL_FACTURA
        WHERE LICITACION IN ({placeholders}) AND TIPO_OC = 'LICITACION'
          AND {_FG}
        GROUP BY LICITACION, CODIGO, DESCRIPCION, DOC_CODE
        ORDER BY LICITACION, CODIGO
    """, body.licitaciones)
    fac_map: dict = {}
    for r in cur.fetchall():
        lic = str(r[0] or "").strip()
        if lic not in fac_map:
            fac_map[lic] = []
        doc = str(r[3] or "").strip()
        fac_map[lic].append({
            "codigo": str(r[1] or "").strip(),
            "descripcion": str(r[2] or "").strip(),
            "tipo": "Nota Crédito" if doc in ("NT", "CM") else ("Guía" if doc == "GF" else "Factura"),
            "venta": round(float(r[4] or 0)),
            "n_docs": int(r[5] or 0),
            "ultima_fecha": str(r[6] or "")[:10],
        })

    conn.close()

    # ── Generar Excel ──
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    green_font = Font(color="10B981", bold=True)
    red_font = Font(color="EF4444", bold=True)
    border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )
    money_fmt = '#,##0'

    # Hoja Resumen
    headers_res = ["Licitación", "RUT", "Cliente", "Adjudicado", "Facturado", "Gap", "Cumpl.%", "Término"]
    for c, h in enumerate(headers_res, 1):
        cell = ws_res.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Necesitamos facturado por licitación
    fac_totals: dict = {}
    for lic, items in fac_map.items():
        fac_totals[lic] = sum(p["venta"] for p in items)

    for i, r in enumerate(resumen, 2):
        fac_total = fac_totals.get(r["licitacion"], 0)
        gap = r["adjudicado"] - fac_total
        pct = round(fac_total / r["adjudicado"] * 100, 1) if r["adjudicado"] > 0 else 0
        ws_res.cell(row=i, column=1, value=r["licitacion"]).border = border
        ws_res.cell(row=i, column=2, value=r["rut"]).border = border
        ws_res.cell(row=i, column=3, value=r["cliente"]).border = border
        c_adj = ws_res.cell(row=i, column=4, value=r["adjudicado"])
        c_adj.number_format = money_fmt
        c_adj.border = border
        c_fac = ws_res.cell(row=i, column=5, value=fac_total)
        c_fac.number_format = money_fmt
        c_fac.font = green_font
        c_fac.border = border
        c_gap = ws_res.cell(row=i, column=6, value=gap)
        c_gap.number_format = money_fmt
        c_gap.font = red_font
        c_gap.border = border
        c_pct = ws_res.cell(row=i, column=7, value=pct)
        c_pct.number_format = '0.0"%"'
        c_pct.alignment = Alignment(horizontal="center")
        c_pct.border = border
        ws_res.cell(row=i, column=8, value=r["fecha_termino"]).border = border

    # Ajustar anchos
    ws_res.column_dimensions["A"].width = 20
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 45
    ws_res.column_dimensions["D"].width = 16
    ws_res.column_dimensions["E"].width = 16
    ws_res.column_dimensions["F"].width = 16
    ws_res.column_dimensions["G"].width = 10
    ws_res.column_dimensions["H"].width = 14

    # ── Hoja Adjudicados ──
    lic_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    lic_font = Font(bold=True, size=11, color="1E40AF")

    ws_adj = wb.create_sheet("Productos Adjudicados")
    h_adj = ["Producto Licitación", "Producto LBF", "Monto Adjudicado", "Estado"]
    for c, h in enumerate(h_adj, 1):
        cell = ws_adj.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c >= 3 else "left")

    row = 2
    for r in resumen:
        lic = r["licitacion"]
        items = adj_map.get(lic, [])
        if not items:
            continue
        # Fila separadora con nombre de licitación + cliente
        ws_adj.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws_adj.cell(row=row, column=1, value=f"{lic} — {r['cliente']}")
        cell.font = lic_font
        cell.fill = lic_fill
        cell.border = border
        row += 1
        for a in items:
            ws_adj.cell(row=row, column=1, value=a["producto"]).border = border
            ws_adj.cell(row=row, column=2, value=a["producto_lbf"]).border = border
            c_m = ws_adj.cell(row=row, column=3, value=a["monto"])
            c_m.number_format = money_fmt
            c_m.border = border
            ws_adj.cell(row=row, column=4, value=a["estado"]).border = border
            row += 1
        row += 1  # Línea en blanco entre licitaciones

    ws_adj.column_dimensions["A"].width = 55
    ws_adj.column_dimensions["B"].width = 55
    ws_adj.column_dimensions["C"].width = 18
    ws_adj.column_dimensions["D"].width = 14

    # ── Hoja Facturados ──
    ws_fac = wb.create_sheet("Productos Facturados")
    h_fac = ["Código", "Descripción", "Tipo Doc", "Facturado", "# Docs", "Última Fecha"]
    for c, h in enumerate(h_fac, 1):
        cell = ws_fac.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c >= 3 else "left")

    row = 2
    for r in resumen:
        lic = r["licitacion"]
        items = fac_map.get(lic, [])
        if not items:
            continue
        # Fila separadora
        ws_fac.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws_fac.cell(row=row, column=1, value=f"{lic} — {r['cliente']}")
        cell.font = lic_font
        cell.fill = lic_fill
        cell.border = border
        row += 1
        for p in items:
            ws_fac.cell(row=row, column=1, value=p["codigo"]).border = border
            ws_fac.cell(row=row, column=2, value=p["descripcion"]).border = border
            ws_fac.cell(row=row, column=3, value=p["tipo"]).border = border
            c_v = ws_fac.cell(row=row, column=4, value=p["venta"])
            c_v.number_format = money_fmt
            c_v.font = green_font
            c_v.border = border
            ws_fac.cell(row=row, column=5, value=p["n_docs"]).border = border
            ws_fac.cell(row=row, column=6, value=p["ultima_fecha"]).border = border
            row += 1
        row += 1

    ws_fac.column_dimensions["A"].width = 16
    ws_fac.column_dimensions["B"].width = 55
    ws_fac.column_dimensions["C"].width = 16
    ws_fac.column_dimensions["D"].width = 18
    ws_fac.column_dimensions["E"].width = 10
    ws_fac.column_dimensions["F"].width = 14

    # Escribir a buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Licitaciones_{body.kam.replace(' ', '_')}_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class NotaBody(BaseModel):
    licitacion: str
    nota: str
    autor: str = ""


@router.post("/nota")
async def save_nota(
    body: NotaBody,
    current_user: dict = Depends(get_current_user),
):
    notas = _load_notas()
    autor = body.autor or current_user.get("display_name", current_user.get("username", ""))
    notas[body.licitacion] = {
        "texto": body.nota,
        "autor": autor,
        "fecha": datetime.date.today().isoformat(),
    }
    _save_notas(notas)
    # Invalidar caches para que la nota aparezca al recargar
    _mem_cache.pop("facturacion_vigentes", None)
    _mem_cache.pop(f"fac_det_{body.licitacion}", None)
    return {"status": "ok", "nota": notas[body.licitacion]}


@router.delete("/nota")
async def delete_nota(
    licitacion: str = Query(...),
    current_user: dict = Depends(get_current_user),
):
    notas = _load_notas()
    notas.pop(licitacion, None)
    _save_notas(notas)
    _mem_cache.pop("facturacion_vigentes", None)
    _mem_cache.pop(f"fac_det_{licitacion}", None)
    return {"status": "ok"}
